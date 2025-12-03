from io import BytesIO
from datetime import date, timedelta
from typing import Iterable, List, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from sqlalchemy.orm import Session
from sqlalchemy import func

from .. import models

# ---------- EXCEL ----------
def _autosize(ws):
    for col in ws.columns:
        maxlen = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            maxlen = max(maxlen, len(val))
        ws.column_dimensions[col_letter].width = min(maxlen + 2, 40)

def excel_products(products: Iterable[models.Product]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Prodotti"

    ws.append(["ID", "Nome", "SKU", "Prezzo (EUR)", "Unità", "Quantità", "Min. scorta", "Attivo"])
    for p in products:
        ws.append([
            p.id,
            p.name,
            p.sku or "",
            (p.price_cents or 0) / 100.0,
            p.unit,
            p.quantity,
            p.min_quantity,
            "Sì" if p.is_active else "No",
        ])

    _autosize(ws)
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()

def products_for_shopping(db: Session, supplier: str | None, only_low_stock: bool, days_to_expiry: int | None):
    q = (
        db.query(
            models.Product,
            func.coalesce(func.sum(models.Lot.quantity), 0.0).label("qty_sum"),
            func.min(models.Lot.expiry_date).label("min_expiry"),
        )
        .outerjoin(models.Lot, models.Lot.product_id == models.Product.id)
        .filter(models.Product.is_active == True)
        .group_by(models.Product.id)
    )
    if supplier:
        q = q.filter(models.Product.supplier == supplier)

    rows = q.all()
    out = []
    from datetime import date, timedelta
    limit_date = date.today() + timedelta(days=days_to_expiry) if days_to_expiry is not None else None

    for prod, qty_sum, min_expiry in rows:
        qty_sum = float(qty_sum or 0)
        low = qty_sum <= float(prod.min_quantity or 0)
        exp_ok = True
        if limit_date is not None:
            exp_ok = (min_expiry is not None and min_expiry <= limit_date)
        if only_low_stock and not low:
            continue
        if days_to_expiry is not None and not exp_ok:
            continue
        # aggiorno quantity per l'export
        prod.quantity = qty_sum
        # uso min_expiry per la colonna Scadenza
        prod.expiry_date = min_expiry
        out.append(prod)
    return out

def excel_closeouts(closeouts: Iterable[models.Closeout]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Chiusure"

    ws.append([
        "ID", "Data", "Cassa Start", "Cassa End", "Vendite contanti", "Vendite POS",
        "Altri incassi", "Spese", "Atteso in cassa", "Scostamento", "Stato", "Note"
    ])

    for cl in closeouts:
        total_income_reported = float(cl.cash_sales) + float(cl.pos_sales) + float(cl.other_income)
        expected_cash_end = float(cl.cash_start) + float(cl.cash_sales) + float(cl.other_income) - float(cl.expenses)
        variance = float(cl.cash_end) - expected_cash_end
        status = "OK"
        if variance > 0.01: status = "SOPRA CASSA"
        elif variance < -0.01: status = "SOTTO CASSA"

        ws.append([
            cl.id, cl.business_date, cl.cash_start, cl.cash_end, cl.cash_sales, cl.pos_sales,
            cl.other_income, cl.expenses, round(expected_cash_end, 2), round(variance, 2), status, cl.note or ""
        ])

    _autosize(ws)
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()

# ---------- PDF (scheda singola closeout) ----------
def pdf_closeout_detail(cl: models.Closeout) -> bytes:
    total_income_reported = float(cl.cash_sales) + float(cl.pos_sales) + float(cl.other_income)
    expected_cash_end = float(cl.cash_start) + float(cl.cash_sales) + float(cl.other_income) - float(cl.expenses)
    variance = float(cl.cash_end) - expected_cash_end
    status = "OK"
    if variance > 0.01: status = "SOPRA CASSA"
    elif variance < -0.01: status = "SOTTO CASSA"

    bio = BytesIO()
    c = canvas.Canvas(bio, pagesize=A4)
    w, h = A4

    # Header
    c.setFont("Helvetica-Bold", 16)
    c.drawString(20*mm, (h - 25*mm), "Gestionale Birreria - Chiusura di cassa")
    c.setFont("Helvetica", 11)
    c.drawString(20*mm, (h - 35*mm), f"Data esercizio: {cl.business_date.isoformat()}")
    c.drawString(20*mm, (h - 42*mm), f"ID chiusura: {cl.id}")

    # Box stato
    if status == "OK": fill = colors.green
    elif status == "SOPRA CASSA": fill = colors.orange
    else: fill = colors.red

    c.setFillColor(fill)
    c.rect(20*mm, (h - 55*mm), 30*mm, 8*mm, stroke=0, fill=1)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 10)
    c.drawCentredString(35*mm, (h - 53*mm), status)

    # Dati principali
    c.setFillColor(colors.black)
    c.setFont("Helvetica", 11)

    rows: List[Tuple[str, str]] = [
        ("Cassa iniziale", f"{cl.cash_start:.2f} €"),
        ("Cassa finale", f"{cl.cash_end:.2f} €"),
        ("Vendite contanti", f"{cl.cash_sales:.2f} €"),
        ("Vendite POS", f"{cl.pos_sales:.2f} €"),
        ("Altri incassi", f"{cl.other_income:.2f} €"),
        ("Spese", f"{cl.expenses:.2f} €"),
        ("Atteso in cassa", f"{expected_cash_end:.2f} €"),
        ("Scostamento", f"{variance:+.2f} €"),
        ("Totale incassi riportati", f"{total_income_reported:.2f} €"),
    ]

    x1, x2 = 20*mm, 90*mm
    y = h - 70*mm
    for label, value in rows:
        c.drawString(x1, y, label)
        c.drawRightString(x2, y, value)
        y -= 7*mm

    # Note
    y -= 5*mm
    c.setFont("Helvetica-Bold", 11)
    c.drawString(20*mm, y, "Note")
    c.setFont("Helvetica", 11)
    y -= 6*mm
    text = cl.note or "-"
    for line in _wrap_text(text, 90):  # 90 chars naive wrap
        c.drawString(20*mm, y, line)
        y -= 6*mm

    c.showPage()
    c.save()
    return bio.getvalue()

def _wrap_text(s: str, width: int) -> List[str]:
    words = (s or "").split()
    if not words: return ["-"]
    lines, cur = [], ""
    for w in words:
        if len(cur) + len(w) + 1 > width:
            lines.append(cur)
            cur = w
        else:
            cur = (w if not cur else f"{cur} {w}")
    if cur: lines.append(cur)
    return lines
